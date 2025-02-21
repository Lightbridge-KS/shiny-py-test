"""Upload Spreadsheet"""

from pathlib import Path
import pandas as pd

from openpyxl import load_workbook
from shiny import App, Inputs, Outputs, Session, reactive, render, ui, req
from shiny.types import FileInfo

def list_sheet_names(excel_path):
    """
    List all sheet names in an Excel file.
    """
    # Load the workbook
    wb = load_workbook(excel_path, read_only=True)
    # Get sheet names
    sheet_names = wb.sheetnames
    # Close the workbook
    wb.close()
    return sheet_names



app_ui = ui.page_fluid(
    ui.input_file("file_input", "Choose Excel or CSV File", accept=[".csv", ".xlsx"], multiple=False),
    ui.input_select(
        "excel_sheet_select",
        "Select Excel Sheet:",
        [],
    ),
    ui.output_data_frame("output_df"),
)

def server(input: Inputs, output: Outputs, session: Session):
    
    @reactive.calc
    def parsed_file():
        file: list[FileInfo] | None = input.file_input()
        if file is None:
            return pd.DataFrame()
        
        file_path =  Path(file[0]["datapath"])
        
        if file_path.suffix.lower() == ".csv":
            return pd.read_csv(file_path)
        
        elif file_path.suffix.lower() == ".xlsx":
            sheet_name = input.excel_sheet_select()
            if len(sheet_name) == 0:
                return pd.DataFrame()
            else:
                return pd.read_excel(file_path, sheet_name = input.excel_sheet_select())
        else:
            raise ValueError(f"Invalid file extension: {file_path.suffix}")
    
    @reactive.calc
    def get_excel_sheet_name():
        file: list[FileInfo] | None = input.file_input()
        req(input.file_input())
        file_path =  Path(file[0]["datapath"])
        if file_path.suffix.lower() == ".xlsx":
            sheet_names = list_sheet_names(file_path)
            return sheet_names
    
    @reactive.Effect
    def update_excel_sheet_select():
        sheet_names = get_excel_sheet_name()
        ui.update_select("excel_sheet_select",
                         choices = sheet_names,
                         )
    
    @render.data_frame
    def output_df():
        df = parsed_file()
        # print(f"type: {type(df)}")
        return render.DataGrid(df)

app = App(app_ui, server)